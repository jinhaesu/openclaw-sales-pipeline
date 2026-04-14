from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .standards import (
    CHANNEL_OUTPUT_CONTRACT_ID,
    EXCEL_POSTPROCESS_RULESET_ID,
    PRODUCT_ANALYSIS_MASTER_SCHEMA_ID,
    build_channel_output_contract,
    build_product_analysis_master_schema,
    merge_postprocess_rules,
)


HEADER_CANDIDATES = {
    "sku": ["SKU", "sku", "상품코드", "품목코드", "품번", "자체상품코드", "판매자상품코드", "옵션관리코드"],
    "product": ["상품명", "품목명", "상품", "품목", "제품명", "옵션명", "상품명(옵션포함)"],
    "qty": ["수량", "판매수량", "주문수량", "구매수량"],
    "sales": ["매출", "매출액", "판매금액", "결제금액", "주문금액", "실결제금액", "상품구매금액"],
    "gross_sales": ["총매출", "주문금액", "주문합계금액", "상품주문금액", "결제예정금액"],
    "net_sales": ["순매출", "순매출액", "실결제금액", "결제금액", "정산금액", "최종결제금액"],
    "supply_amount": ["공급가", "공급금액", "공급가액", "입고금액"],
    "delivery_amount": ["납품금액", "납입금액", "배송금액"],
    "date": ["일자", "날짜", "주문일", "판매일", "결제일", "정산일", "출고일", "등록일", "작성일"],
    "status": ["상태", "주문상태", "배송상태", "처리상태", "진행상태", "구분"],
    "refund_flag": ["환불여부", "취소여부", "반품여부"],
}


def analyze_sales_file(
    path: Path,
    profile: dict[str, Any] | None = None,
    context: dict[str, Any] | None = None,
) -> dict[str, Any]:
    profile = profile or {}
    context = context or {}
    rows = load_rows(path)
    header_index, headers = detect_header(rows)
    records = rows_to_dicts(rows[header_index + 1 :], headers)
    applied_rules = merge_postprocess_rules(profile)
    normalized, quality = normalize_records(
        records,
        profile=profile,
        context=context,
        source_path=path,
        applied_rules=applied_rules,
    )
    summary = summarize_by_product(normalized)
    totals = summarize_totals(normalized)
    channel_summary = {
        "vendor_name": context.get("vendor_name", ""),
        "business_date": context.get("business_date", ""),
        "business_month": (context.get("business_date", "") or "")[:7],
        "revenue_basis": context.get("revenue_basis", ""),
        "date_basis": context.get("date_basis", ""),
        "row_count": len(normalized),
        "product_count": len(summary),
        "total_qty": totals["qty"],
        "total_sales": totals["sales"],
        "total_gross_sales": totals["gross_sales"],
        "total_net_sales": totals["net_sales"],
        "total_supply_amount": totals["supply_amount"],
        "total_delivery_amount": totals["delivery_amount"],
        "source_file": str(path),
    }
    return {
        "output_type": "channel_sales_analysis",
        "format_version": "2026-04-14",
        "channel_output_contract_id": CHANNEL_OUTPUT_CONTRACT_ID,
        "product_analysis_master_schema_id": PRODUCT_ANALYSIS_MASTER_SCHEMA_ID,
        "excel_postprocess_ruleset_id": EXCEL_POSTPROCESS_RULESET_ID,
        "metadata": {
            "vendor_name": context.get("vendor_name", ""),
            "channel_group": context.get("channel_group", ""),
            "manager": context.get("manager", ""),
            "business_date": context.get("business_date", ""),
            "revenue_basis": context.get("revenue_basis", ""),
            "date_basis": context.get("date_basis", ""),
            "revenue_metric_key": context.get("revenue_metric_key", ""),
            "source_file": str(path),
        },
        "channel_summary": channel_summary,
        "applied_postprocess_rules": applied_rules,
        "quality": {
            "header_row_index": header_index,
            "detected_headers": headers,
            **quality,
        },
        "schemas": {
            "channel_output_contract": build_channel_output_contract(),
            "product_analysis_master_schema": build_product_analysis_master_schema(),
        },
        "file": str(path),
        "vendor_name": context.get("vendor_name"),
        "business_date": context.get("business_date"),
        "row_count": len(normalized),
        "product_count": len(summary),
        "totals": totals,
        "items": summary,
        "records": normalized,
    }


def load_rows(path: Path) -> list[list[Any]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".txt"}:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [row for row in csv.reader(handle)]
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, data_only=True, read_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    raise ValueError(f"Unsupported file type: {path.suffix}")


def detect_header(rows: list[list[Any]]) -> tuple[int, list[str]]:
    for idx, row in enumerate(rows[:20]):
        text_row = [str(cell).strip() if cell is not None else "" for cell in row]
        if score_header(text_row) >= 2:
            return idx, text_row
    fallback = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    return 0, fallback


def score_header(row: list[str]) -> int:
    score = 0
    for candidates in HEADER_CANDIDATES.values():
        if any(cell in candidates for cell in row):
            score += 1
    return score


def rows_to_dicts(rows: list[list[Any]], headers: list[str]) -> list[dict[str, Any]]:
    result = []
    width = len(headers)
    for row in rows:
        padded = list(row[:width]) + [None] * max(0, width - len(row))
        result.append({headers[i] or f"col_{i}": padded[i] for i in range(width)})
    return result


def normalize_records(
    records: list[dict[str, Any]],
    profile: dict[str, Any] | None = None,
    context: dict[str, Any] | None = None,
    source_path: Path | None = None,
    applied_rules: dict[str, Any] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    profile = profile or {}
    context = context or {}
    header_candidates = build_header_candidates(profile)
    applied_rules = applied_rules or merge_postprocess_rules(profile)
    exclude_status_keywords = [str(item).strip().lower() for item in applied_rules.get("exclude_status_keywords", [])]
    include_status_keywords = [str(item).strip().lower() for item in applied_rules.get("include_status_keywords", [])]
    exclude_product_keywords = [str(item).strip().lower() for item in applied_rules.get("exclude_product_keywords", [])]
    footer_keywords = [str(item).strip().lower() for item in applied_rules.get("drop_footer_keywords", [])]
    normalized = []
    stats = {
        "input_row_count": len(records),
        "kept_row_count": 0,
        "dropped_empty_rows": 0,
        "dropped_footer_rows": 0,
        "dropped_status_rows": 0,
        "dropped_product_rows": 0,
        "missing_date_rows": 0,
    }
    for record in records:
        sku = normalize_text(first_value(record, header_candidates["sku"]))
        product = first_value(record, header_candidates["product"])
        qty = to_number(first_value(record, header_candidates["qty"]))
        raw_sales = to_number(first_value(record, header_candidates["sales"]))
        gross_sales = to_number(first_value(record, header_candidates["gross_sales"]))
        net_sales = to_number(first_value(record, header_candidates["net_sales"]))
        supply_amount = to_number(first_value(record, header_candidates["supply_amount"]))
        delivery_amount = to_number(first_value(record, header_candidates["delivery_amount"]))
        status = normalize_text(first_value(record, header_candidates["status"]))
        business_date = to_iso_date(first_value(record, header_candidates["date"])) or context.get("business_date")
        if not product and qty == 0 and raw_sales == 0 and gross_sales == 0 and net_sales == 0 and supply_amount == 0 and delivery_amount == 0:
            stats["dropped_empty_rows"] += 1
            continue
        product_name = str(product).strip() if product is not None else "(unknown)"
        if footer_keywords and contains_keyword(product_name, footer_keywords):
            stats["dropped_footer_rows"] += 1
            continue
        if exclude_status_keywords and contains_keyword(status, exclude_status_keywords):
            stats["dropped_status_rows"] += 1
            continue
        if include_status_keywords and not contains_keyword(status, include_status_keywords):
            stats["dropped_status_rows"] += 1
            continue
        if exclude_product_keywords and contains_keyword(product_name, exclude_product_keywords):
            stats["dropped_product_rows"] += 1
            continue
        normalized_product_name = normalize_product_name(product_name, applied_rules)
        if not normalized_product_name:
            stats["dropped_product_rows"] += 1
            continue
        if not business_date:
            stats["missing_date_rows"] += 1
        metric_key = str(context.get("revenue_metric_key", "") or "").strip()
        sales = resolve_primary_amount(
            metric_key,
            {
                "sales": raw_sales,
                "gross_sales": gross_sales,
                "net_sales": net_sales,
                "supply_amount": supply_amount,
                "delivery_amount": delivery_amount,
            },
            applied_rules,
        )
        refund_flag = infer_refund_flag(record, status, product_name, header_candidates, applied_rules)
        normalized.append(
            {
                "vendor_name": context.get("vendor_name", ""),
                "channel_group": context.get("channel_group", ""),
                "manager": context.get("manager", ""),
                "business_date": business_date or "",
                "business_month": business_date[:7] if business_date else "",
                "status": status,
                "sku": sku,
                "product_name": product_name,
                "normalized_product_name": normalized_product_name,
                "qty": qty,
                "sales": sales,
                "gross_sales": gross_sales,
                "net_sales": net_sales,
                "supply_amount": supply_amount,
                "delivery_amount": delivery_amount,
                "refund_flag": refund_flag,
                "revenue_basis": context.get("revenue_basis", ""),
                "date_basis": context.get("date_basis", ""),
                "queue_id": context.get("queue_id", ""),
                "queue_label": context.get("queue_label", ""),
                "collection_mode": context.get("collection_mode", ""),
                "browser_policy": context.get("browser_policy", ""),
                "session_strategy": context.get("session_strategy", ""),
                "validation_mode": context.get("validation_mode", ""),
                "verification_mode": context.get("verification_mode", ""),
                "source_file": str(source_path) if source_path else "",
                "raw_file_name": source_path.name if source_path else "",
                "raw": record,
            }
        )
        stats["kept_row_count"] += 1
    return normalized, stats


def first_value(record: dict[str, Any], candidates: list[str]) -> Any:
    for candidate in candidates:
        if candidate in record and record[candidate] not in (None, ""):
            return record[candidate]
    return None


def to_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def build_header_candidates(profile: dict[str, Any]) -> dict[str, list[str]]:
    candidates = {key: list(values) for key, values in HEADER_CANDIDATES.items()}
    overrides = {
        "sku": profile.get("sku_header_candidates", []),
        "product": profile.get("product_header_candidates", []),
        "qty": profile.get("qty_header_candidates", []),
        "sales": profile.get("sales_header_candidates", []),
        "gross_sales": profile.get("gross_sales_header_candidates", []),
        "net_sales": profile.get("net_sales_header_candidates", []),
        "supply_amount": profile.get("supply_amount_header_candidates", []),
        "delivery_amount": profile.get("delivery_amount_header_candidates", []),
        "date": profile.get("date_header_candidates", []),
        "status": profile.get("status_header_candidates", []),
        "refund_flag": profile.get("refund_flag_header_candidates", []),
    }
    for key, items in overrides.items():
        for item in items:
            if item not in candidates[key]:
                candidates[key].insert(0, item)
    return candidates


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def contains_keyword(value: str, keywords: list[str]) -> bool:
    lowered = normalize_text(value).lower()
    return any(keyword and keyword in lowered for keyword in keywords)


def normalize_product_name(value: str, rules: dict[str, Any]) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    aliases = {str(key): str(val) for key, val in rules.get("product_name_aliases", {}).items()}
    if text in aliases:
        text = aliases[text]
    if rules.get("strip_bracket_suffixes"):
        text = re.sub(r"\s*[\(\[].*?[\)\]]\s*$", "", text)
    if rules.get("normalize_whitespace", True):
        text = re.sub(r"\s+", " ", text).strip()
    return text


def resolve_primary_amount(
    preferred_metric_key: str,
    metric_values: dict[str, float],
    rules: dict[str, Any],
) -> float:
    preferred = metric_values.get(preferred_metric_key, 0.0) if preferred_metric_key else 0.0
    if preferred:
        return round(preferred, 2)
    for key in rules.get("amount_priority_columns", []):
        value = metric_values.get(str(key), 0.0)
        if value:
            return round(value, 2)
    return round(metric_values.get("sales", 0.0), 2)


def infer_refund_flag(
    record: dict[str, Any],
    status: str,
    product_name: str,
    header_candidates: dict[str, list[str]],
    rules: dict[str, Any],
) -> bool:
    explicit_value = normalize_text(first_value(record, header_candidates["refund_flag"]))
    if explicit_value in {"y", "yes", "true", "1", "환불", "취소", "반품"}:
        return True
    refund_keywords = [str(item).strip().lower() for item in rules.get("refund_status_keywords", [])]
    if contains_keyword(status, refund_keywords):
        return True
    if contains_keyword(product_name, refund_keywords):
        return True
    return False


def to_iso_date(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = normalize_text(value)
    if not text:
        return ""
    text = text.replace(".", "-").replace("/", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m", "%Y%m%d", "%Y%m", "%m-%d-%Y", "%m-%d-%y"):
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt in {"%Y-%m", "%Y%m"}:
                return parsed.strftime("%Y-%m-01")
            return parsed.date().isoformat()
        except ValueError:
            continue
    return ""


def summarize_by_product(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    bucket: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "qty": 0.0,
            "sales": 0.0,
            "gross_sales": 0.0,
            "net_sales": 0.0,
            "supply_amount": 0.0,
            "delivery_amount": 0.0,
            "sku": "",
            "display_name": "",
        }
    )
    for record in records:
        key = record.get("normalized_product_name") or record["product_name"]
        bucket[key]["qty"] += record["qty"]
        bucket[key]["sales"] += record["sales"]
        bucket[key]["gross_sales"] += float(record.get("gross_sales", 0) or 0)
        bucket[key]["net_sales"] += float(record.get("net_sales", 0) or 0)
        bucket[key]["supply_amount"] += float(record.get("supply_amount", 0) or 0)
        bucket[key]["delivery_amount"] += float(record.get("delivery_amount", 0) or 0)
        if not bucket[key]["sku"] and record.get("sku"):
            bucket[key]["sku"] = str(record.get("sku", ""))
        if not bucket[key]["display_name"]:
            bucket[key]["display_name"] = record["product_name"]
    items = [
        {
            "sku": value["sku"],
            "product_name": value["display_name"] or key,
            "normalized_product_name": key,
            "qty": round(value["qty"], 2),
            "sales": round(value["sales"], 2),
            "gross_sales": round(value["gross_sales"], 2),
            "net_sales": round(value["net_sales"], 2),
            "supply_amount": round(value["supply_amount"], 2),
            "delivery_amount": round(value["delivery_amount"], 2),
        }
        for key, value in bucket.items()
    ]
    items.sort(key=lambda item: (-item["sales"], item["product_name"]))
    return items


def summarize_totals(records: list[dict[str, Any]]) -> dict[str, float]:
    return {
        "qty": round(sum(record["qty"] for record in records), 2),
        "sales": round(sum(record["sales"] for record in records), 2),
        "gross_sales": round(sum(float(record.get("gross_sales", 0) or 0) for record in records), 2),
        "net_sales": round(sum(float(record.get("net_sales", 0) or 0) for record in records), 2),
        "supply_amount": round(sum(float(record.get("supply_amount", 0) or 0) for record in records), 2),
        "delivery_amount": round(sum(float(record.get("delivery_amount", 0) or 0) for record in records), 2),
    }


def write_analysis(output_path: Path, analysis: dict[str, Any]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(analysis, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
