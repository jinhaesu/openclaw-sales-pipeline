from __future__ import annotations

import base64
import json
import mimetypes
import smtplib
import time
from collections import defaultdict
from dataclasses import asdict
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
from typing import Any
from urllib import error as urlerror
from urllib import request as urlrequest

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .excel_analysis import analyze_sales_file
from .models import ChannelRecord, Playbook
from .operations import build_channel_operation_profile
from .secrets import SecretStore
from .standards import (
    CHANNEL_OUTPUT_CONTRACT_ID,
    EXCEL_POSTPROCESS_RULESET_ID,
    PRODUCT_ANALYSIS_MASTER_SCHEMA_ID,
    build_standards_bundle,
)


SUPPORTED_SOURCE_SUFFIXES = {".xlsx", ".xlsm", ".csv", ".txt"}


def safe_vendor_name(value: str) -> str:
    return value.replace(" ", "_").replace("/", "_").replace("(", "").replace(")", "")


def build_report_bundle(
    channels: list[ChannelRecord],
    playbooks: dict[str, Playbook],
    secrets: SecretStore,
    input_root: Path | None,
    output_dir: Path,
    date_from: str | None = None,
    date_to: str | None = None,
    channel_filters: list[str] | None = None,
    explicit_files: list[str] | None = None,
    manifest_path: str | None = None,
    label: str | None = None,
    email_to: list[str] | None = None,
    email_cc: list[str] | None = None,
    email_subject: str | None = None,
    send_email: bool = False,
    smtp_profile: str = "smtp",
) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    sources = collect_report_sources(
        channels=channels,
        playbooks=playbooks,
        input_root=input_root,
        date_from=date_from,
        date_to=date_to,
        channel_filters=channel_filters or [],
        explicit_files=explicit_files or [],
        manifest_path=manifest_path,
    )
    analyses = analyze_sources(sources, channels, playbooks)
    records = [record for analysis in analyses for record in analysis["records"]]
    aggregates = aggregate_records(records)

    report_label = label or infer_label(analyses, date_from=date_from, date_to=date_to)
    workbook_path = output_dir / f"sales_report_{report_label}.xlsx"
    summary_path = output_dir / f"sales_report_{report_label}.md"
    summary_json_path = output_dir / f"sales_report_{report_label}.json"
    manifest_output_path = output_dir / f"sales_report_{report_label}_manifest.json"

    report = {
        "label": report_label,
        "standards": build_standards_bundle(),
        "source_count": len(sources),
        "analysis_count": len(analyses),
        "record_count": len(records),
        "sources": sources,
        "analyses": analyses,
        "aggregates": aggregates,
        "summary": build_summary(aggregates, analyses),
    }

    export_report_workbook(workbook_path, report)
    summary_markdown = build_summary_markdown(report)
    summary_path.write_text(summary_markdown, encoding="utf-8")
    summary_json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    manifest_output_path.write_text(json.dumps({"sources": sources}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    email_profile_name, email_config = resolve_email_profile(secrets, smtp_profile)

    draft_path = create_email_draft(
        output_dir=output_dir,
        subject=email_subject or f"[OpenClaw] 매출 리포트 {report_label}",
        to_addrs=email_to or [],
        cc_addrs=email_cc or [],
        summary_markdown=summary_markdown,
        attachments=[workbook_path, summary_path],
        from_addr=email_config.get("from_addr", ""),
    )

    smtp_status = validate_smtp_profile(secrets, smtp_profile)
    sent = False
    send_error = ""
    if send_email:
        sent, send_error = send_email_bundle(
            secrets=secrets,
            smtp_profile=smtp_profile,
            subject=email_subject or f"[OpenClaw] 매출 리포트 {report_label}",
            to_addrs=email_to or [],
            cc_addrs=email_cc or [],
            summary_markdown=summary_markdown,
            attachments=[workbook_path, summary_path],
        )

    return {
        "label": report_label,
        "source_count": len(sources),
        "analysis_count": len(analyses),
        "record_count": len(records),
        "outputs": {
            "workbook": str(workbook_path.resolve()),
            "summary_markdown": str(summary_path.resolve()),
            "summary_json": str(summary_json_path.resolve()),
            "manifest": str(manifest_output_path.resolve()),
            "email_draft": str(draft_path.resolve()),
        },
        "standards": {
            "channel_output_contract_id": CHANNEL_OUTPUT_CONTRACT_ID,
            "excel_postprocess_ruleset_id": EXCEL_POSTPROCESS_RULESET_ID,
            "product_analysis_master_schema_id": PRODUCT_ANALYSIS_MASTER_SCHEMA_ID,
        },
        "smtp_status": smtp_status,
        "email_profile": email_profile_name,
        "email_provider": smtp_status.get("provider", ""),
        "sent_email": sent,
        "send_error": send_error,
        "summary": report["summary"],
    }


def collect_report_sources(
    channels: list[ChannelRecord],
    playbooks: dict[str, Playbook],
    input_root: Path | None,
    date_from: str | None,
    date_to: str | None,
    channel_filters: list[str],
    explicit_files: list[str],
    manifest_path: str | None,
) -> list[dict[str, Any]]:
    sources: list[dict[str, Any]] = []
    vendor_lookup = {safe_vendor_name(channel.vendor_name): channel.vendor_name for channel in channels}
    channel_set = set(channel_filters)

    if manifest_path:
        manifest_raw = json.loads(Path(manifest_path).expanduser().read_text(encoding="utf-8"))
        for item in manifest_raw.get("sources", []):
            vendor_name = str(item.get("vendor_name", ""))
            if channel_set and vendor_name not in channel_set:
                continue
            sources.append(
                {
                    "vendor_name": vendor_name,
                    "business_date": str(item.get("business_date", "")),
                    "path": str(Path(item["path"]).expanduser()),
                    "source_type": item.get("source_type", "raw_file"),
                }
            )

    for file_path in explicit_files:
        candidate = Path(file_path).expanduser().resolve()
        inferred = infer_context_from_path(candidate, vendor_lookup)
        vendor_name = inferred.get("vendor_name", "")
        if channel_set and vendor_name and vendor_name not in channel_set:
            continue
        sources.append(
            {
                "vendor_name": vendor_name,
                "business_date": inferred.get("business_date", ""),
                "path": str(candidate),
                "source_type": inferred.get("source_type", "raw_file"),
            }
        )

    if input_root and input_root.exists():
        for date_dir in sorted(path for path in input_root.iterdir() if path.is_dir()):
            business_date = date_dir.name
            if not looks_like_date(business_date):
                continue
            if date_from and business_date < date_from:
                continue
            if date_to and business_date > date_to:
                continue
            for vendor_dir in sorted(path for path in date_dir.iterdir() if path.is_dir()):
                vendor_name = vendor_lookup.get(vendor_dir.name, vendor_dir.name.replace("_", " "))
                if channel_set and vendor_name not in channel_set:
                    continue
                analysis_by_stem: dict[str, Path] = {}
                raw_files: list[Path] = []
                for file_path in sorted(vendor_dir.rglob("*")):
                    if not file_path.is_file():
                        continue
                    suffix = file_path.suffix.lower()
                    if file_path.name.endswith("_analysis.json") or file_path.name == "file_analysis.json":
                        stem = file_path.name.removesuffix("_analysis.json") if file_path.name.endswith("_analysis.json") else file_path.stem
                        analysis_by_stem[stem] = file_path
                    elif suffix in SUPPORTED_SOURCE_SUFFIXES:
                        raw_files.append(file_path)

                for file_path in sorted(analysis_by_stem.values()):
                    sources.append(
                        {
                            "vendor_name": vendor_name,
                            "business_date": business_date,
                            "path": str(file_path.resolve()),
                            "source_type": "analysis_json",
                        }
                    )

                for file_path in sorted(raw_files):
                    if file_path.stem in analysis_by_stem:
                        continue
                    sources.append(
                        {
                            "vendor_name": vendor_name,
                            "business_date": business_date,
                            "path": str(file_path.resolve()),
                            "source_type": "raw_file",
                        }
                    )

    unique: dict[tuple[str, str, str], dict[str, Any]] = {}
    for item in sources:
        key = (item["vendor_name"], item["business_date"], item["path"])
        unique[key] = item
    return sorted(unique.values(), key=lambda item: (item["business_date"], item["vendor_name"], item["path"]))


def infer_context_from_path(path: Path, vendor_lookup: dict[str, str]) -> dict[str, str]:
    parts = list(path.parts)
    for index, part in enumerate(parts[:-1]):
        if looks_like_date(part):
            vendor_part = parts[index + 1] if index + 1 < len(parts) else ""
            return {
                "business_date": part,
                "vendor_name": vendor_lookup.get(vendor_part, vendor_part.replace("_", " ")),
                "source_type": "analysis_json" if path.name.endswith("_analysis.json") else "raw_file",
            }
    return {"business_date": "", "vendor_name": "", "source_type": "raw_file"}


def analyze_sources(
    sources: list[dict[str, Any]],
    channels: list[ChannelRecord],
    playbooks: dict[str, Playbook],
) -> list[dict[str, Any]]:
    channel_lookup = {channel.vendor_name: channel for channel in channels}
    analyses: list[dict[str, Any]] = []
    for source in sources:
        path = Path(source["path"]).expanduser()
        vendor_name = source.get("vendor_name", "")
        channel = channel_lookup.get(vendor_name)
        playbook = playbooks.get(vendor_name)
        operation_profile = build_channel_operation_profile(asdict(channel), asdict(playbook) if playbook else {}) if channel else {}
        context = {
            "vendor_name": vendor_name,
            "business_date": source.get("business_date", ""),
            "channel_group": channel.channel_group if channel else "",
            "manager": channel.manager if channel else "",
            "queue_id": operation_profile.get("queue_id", ""),
            "queue_label": operation_profile.get("queue_label", ""),
            "collection_mode": operation_profile.get("collection_mode", ""),
            "browser_policy": operation_profile.get("browser_policy", ""),
            "session_strategy": operation_profile.get("session_strategy", ""),
            "revenue_basis": operation_profile.get("revenue_basis", ""),
            "revenue_metric_key": operation_profile.get("revenue_metric_key", ""),
            "date_basis": operation_profile.get("date_basis", ""),
            "validation_mode": operation_profile.get("validation_mode", ""),
            "verification_mode": operation_profile.get("verification_mode", ""),
        }
        if source.get("source_type") == "analysis_json":
            raw = json.loads(path.read_text(encoding="utf-8"))
            records = []
            for item in raw.get("records", []):
                record = dict(item)
                if not record.get("vendor_name"):
                    record["vendor_name"] = vendor_name
                if not record.get("business_date"):
                    record["business_date"] = source.get("business_date", "")
                if not record.get("business_month") and record.get("business_date"):
                    record["business_month"] = str(record["business_date"])[:7]
                if not record.get("normalized_product_name"):
                    record["normalized_product_name"] = record.get("product_name", "(unknown)")
                if not record.get("revenue_basis"):
                    record["revenue_basis"] = context.get("revenue_basis", "")
                if not record.get("date_basis"):
                    record["date_basis"] = context.get("date_basis", "")
                if not record.get("queue_id"):
                    record["queue_id"] = context.get("queue_id", "")
                if not record.get("queue_label"):
                    record["queue_label"] = context.get("queue_label", "")
                if not record.get("collection_mode"):
                    record["collection_mode"] = context.get("collection_mode", "")
                if not record.get("browser_policy"):
                    record["browser_policy"] = context.get("browser_policy", "")
                if not record.get("session_strategy"):
                    record["session_strategy"] = context.get("session_strategy", "")
                if not record.get("validation_mode"):
                    record["validation_mode"] = context.get("validation_mode", "")
                if not record.get("verification_mode"):
                    record["verification_mode"] = context.get("verification_mode", "")
                if not record.get("raw_file_name"):
                    record["raw_file_name"] = path.name
                records.append(record)
            if not records and raw.get("items"):
                for item in raw["items"]:
                    records.append(
                        {
                            "vendor_name": vendor_name,
                            "channel_group": context["channel_group"],
                            "manager": context["manager"],
                            "business_date": source.get("business_date", ""),
                            "business_month": source.get("business_date", "")[:7] if source.get("business_date") else "",
                            "status": "",
                            "product_name": item.get("product_name", "(unknown)"),
                            "normalized_product_name": item.get("normalized_product_name", item.get("product_name", "(unknown)")),
                            "sku": item.get("sku", ""),
                            "qty": float(item.get("qty", 0) or 0),
                            "sales": float(item.get("sales", 0) or 0),
                            "gross_sales": float(item.get("gross_sales", 0) or 0),
                            "net_sales": float(item.get("net_sales", 0) or 0),
                            "supply_amount": float(item.get("supply_amount", 0) or 0),
                            "delivery_amount": float(item.get("delivery_amount", 0) or 0),
                            "refund_flag": bool(item.get("refund_flag", False)),
                            "revenue_basis": context.get("revenue_basis", ""),
                            "date_basis": context.get("date_basis", ""),
                            "queue_id": context.get("queue_id", ""),
                            "queue_label": context.get("queue_label", ""),
                            "collection_mode": context.get("collection_mode", ""),
                            "browser_policy": context.get("browser_policy", ""),
                            "session_strategy": context.get("session_strategy", ""),
                            "validation_mode": context.get("validation_mode", ""),
                            "verification_mode": context.get("verification_mode", ""),
                            "source_file": str(path),
                            "raw_file_name": path.name,
                        }
                    )
            analysis = {
                "output_type": raw.get("output_type", "channel_sales_analysis"),
                "format_version": raw.get("format_version", "2026-04-14"),
                "channel_output_contract_id": raw.get("channel_output_contract_id", CHANNEL_OUTPUT_CONTRACT_ID),
                "product_analysis_master_schema_id": raw.get(
                    "product_analysis_master_schema_id",
                    PRODUCT_ANALYSIS_MASTER_SCHEMA_ID,
                ),
                "excel_postprocess_ruleset_id": raw.get(
                    "excel_postprocess_ruleset_id",
                    EXCEL_POSTPROCESS_RULESET_ID,
                ),
                "metadata": raw.get("metadata", {}),
                "channel_summary": raw.get("channel_summary", {}),
                "applied_postprocess_rules": raw.get("applied_postprocess_rules", {}),
                "quality": raw.get("quality", {}),
                "vendor_name": vendor_name,
                "business_date": source.get("business_date", ""),
                "file": str(path),
                "operation_profile": operation_profile,
                "row_count": len(records),
                "product_count": len({item.get("product_name", "") for item in records}),
                "totals": {
                    "qty": round(sum(float(item.get("qty", 0) or 0) for item in records), 2),
                    "sales": round(sum(float(item.get("sales", 0) or 0) for item in records), 2),
                    "gross_sales": round(sum(float(item.get("gross_sales", 0) or 0) for item in records), 2),
                    "net_sales": round(sum(float(item.get("net_sales", 0) or 0) for item in records), 2),
                    "supply_amount": round(sum(float(item.get("supply_amount", 0) or 0) for item in records), 2),
                    "delivery_amount": round(sum(float(item.get("delivery_amount", 0) or 0) for item in records), 2),
                },
                "records": records,
            }
        else:
            analysis = analyze_sales_file(
                path=path,
                profile=build_analysis_profile(playbook),
                context=context,
            )
            analysis["operation_profile"] = operation_profile
        analyses.append(analysis)
    return analyses


def build_analysis_profile(playbook: Playbook | None) -> dict[str, Any]:
    if not playbook:
        return {}
    profile = dict(playbook.analysis_profile)
    if playbook.postprocess_rules:
        profile["postprocess_rules"] = dict(playbook.postprocess_rules)
    return profile


def aggregate_records(records: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    daily_channel: dict[tuple[str, str], dict[str, Any]] = defaultdict(lambda: _new_amount_bucket())
    monthly_channel: dict[tuple[str, str], dict[str, Any]] = defaultdict(lambda: _new_amount_bucket())
    channel_totals: dict[str, dict[str, Any]] = defaultdict(lambda: _new_amount_bucket())
    product_totals: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "qty": 0.0,
            "sales": 0.0,
            "gross_sales": 0.0,
            "net_sales": 0.0,
            "supply_amount": 0.0,
            "delivery_amount": 0.0,
            "channels": set(),
            "display_name": "",
            "sku": "",
        }
    )
    channel_product: dict[tuple[str, str], dict[str, Any]] = defaultdict(lambda: _new_amount_bucket())
    daily_product: dict[tuple[str, str], dict[str, Any]] = defaultdict(lambda: _new_amount_bucket())
    product_names: dict[str, str] = {}
    channel_metadata: dict[str, dict[str, Any]] = {}

    for record in records:
        vendor_name = str(record.get("vendor_name", ""))
        business_date = str(record.get("business_date", ""))
        business_month = str(record.get("business_month", ""))
        product_key = str(record.get("normalized_product_name") or record.get("product_name", "(unknown)"))
        product_name = str(record.get("product_name", product_key))
        sku = str(record.get("sku", "") or "")
        qty = float(record.get("qty", 0) or 0)
        sales = float(record.get("sales", 0) or 0)
        gross_sales = float(record.get("gross_sales", 0) or 0)
        net_sales = float(record.get("net_sales", 0) or 0)
        supply_amount = float(record.get("supply_amount", 0) or 0)
        delivery_amount = float(record.get("delivery_amount", 0) or 0)
        revenue_basis = str(record.get("revenue_basis", "") or "")
        date_basis = str(record.get("date_basis", "") or "")

        if vendor_name:
            meta = channel_metadata.setdefault(
                vendor_name,
                {
                    "vendor_name": vendor_name,
                    "revenue_basis": revenue_basis,
                    "date_basis": date_basis,
                    "collection_mode": str(record.get("collection_mode", "") or ""),
                    "validation_mode": str(record.get("validation_mode", "") or ""),
                    "queue_id": str(record.get("queue_id", "") or ""),
                    "browser_policy": str(record.get("browser_policy", "") or ""),
                    "verification_mode": str(record.get("verification_mode", "") or ""),
                },
            )
            if not meta["revenue_basis"] and revenue_basis:
                meta["revenue_basis"] = revenue_basis
            if not meta["date_basis"] and date_basis:
                meta["date_basis"] = date_basis

        if business_date:
            update_amount_bucket(daily_channel[(business_date, vendor_name)], qty, sales, gross_sales, net_sales, supply_amount, delivery_amount, revenue_basis, date_basis)
            update_amount_bucket(daily_product[(business_date, product_key)], qty, sales, gross_sales, net_sales, supply_amount, delivery_amount, revenue_basis, date_basis)
        if business_month:
            update_amount_bucket(monthly_channel[(business_month, vendor_name)], qty, sales, gross_sales, net_sales, supply_amount, delivery_amount, revenue_basis, date_basis)
        update_amount_bucket(channel_totals[vendor_name], qty, sales, gross_sales, net_sales, supply_amount, delivery_amount, revenue_basis, date_basis)
        product_totals[product_key]["qty"] += qty
        product_totals[product_key]["sales"] += sales
        product_totals[product_key]["gross_sales"] += gross_sales
        product_totals[product_key]["net_sales"] += net_sales
        product_totals[product_key]["supply_amount"] += supply_amount
        product_totals[product_key]["delivery_amount"] += delivery_amount
        product_totals[product_key]["channels"].add(vendor_name)
        if not product_totals[product_key]["sku"] and sku:
            product_totals[product_key]["sku"] = sku
        if not product_totals[product_key]["display_name"]:
            product_totals[product_key]["display_name"] = product_name
        product_names[product_key] = product_totals[product_key]["display_name"] or product_name
        update_amount_bucket(channel_product[(vendor_name, product_key)], qty, sales, gross_sales, net_sales, supply_amount, delivery_amount, revenue_basis, date_basis)

    return {
        "daily_channel_sales": sort_rows(
            [
                {
                    "business_date": key[0],
                    "vendor_name": key[1],
                    "revenue_basis": value["revenue_basis"],
                    "date_basis": value["date_basis"],
                    "qty": round(value["qty"], 2),
                    "sales": round(value["sales"], 2),
                    "gross_sales": round(value["gross_sales"], 2),
                    "net_sales": round(value["net_sales"], 2),
                    "supply_amount": round(value["supply_amount"], 2),
                    "delivery_amount": round(value["delivery_amount"], 2),
                }
                for key, value in daily_channel.items()
            ],
            ["business_date", "vendor_name"],
        ),
        "monthly_channel_sales": sort_rows(
            [
                {
                    "business_month": key[0],
                    "vendor_name": key[1],
                    "revenue_basis": value["revenue_basis"],
                    "date_basis": value["date_basis"],
                    "qty": round(value["qty"], 2),
                    "sales": round(value["sales"], 2),
                    "gross_sales": round(value["gross_sales"], 2),
                    "net_sales": round(value["net_sales"], 2),
                    "supply_amount": round(value["supply_amount"], 2),
                    "delivery_amount": round(value["delivery_amount"], 2),
                }
                for key, value in monthly_channel.items()
            ],
            ["business_month", "vendor_name"],
        ),
        "channel_totals": sorted(
            [
                {
                    "vendor_name": vendor_name,
                    "revenue_basis": channel_metadata.get(vendor_name, {}).get("revenue_basis", value["revenue_basis"]),
                    "date_basis": channel_metadata.get(vendor_name, {}).get("date_basis", value["date_basis"]),
                    "collection_mode": channel_metadata.get(vendor_name, {}).get("collection_mode", ""),
                    "validation_mode": channel_metadata.get(vendor_name, {}).get("validation_mode", ""),
                    "queue_id": channel_metadata.get(vendor_name, {}).get("queue_id", ""),
                    "browser_policy": channel_metadata.get(vendor_name, {}).get("browser_policy", ""),
                    "verification_mode": channel_metadata.get(vendor_name, {}).get("verification_mode", ""),
                    "qty": round(value["qty"], 2),
                    "sales": round(value["sales"], 2),
                    "gross_sales": round(value["gross_sales"], 2),
                    "net_sales": round(value["net_sales"], 2),
                    "supply_amount": round(value["supply_amount"], 2),
                    "delivery_amount": round(value["delivery_amount"], 2),
                }
                for vendor_name, value in channel_totals.items()
            ],
            key=lambda item: (-item["sales"], item["vendor_name"]),
        ),
        "product_sales": sorted(
            [
                {
                    "sku": value["sku"],
                    "product_name": value["display_name"] or key,
                    "normalized_product_name": key,
                    "sales": round(value["sales"], 2),
                    "qty": round(value["qty"], 2),
                    "gross_sales": round(value["gross_sales"], 2),
                    "net_sales": round(value["net_sales"], 2),
                    "supply_amount": round(value["supply_amount"], 2),
                    "delivery_amount": round(value["delivery_amount"], 2),
                    "channel_count": len(value["channels"]),
                }
                for key, value in product_totals.items()
            ],
            key=lambda item: (-item["sales"], item["product_name"]),
        ),
        "product_qty": sorted(
            [
                {
                    "sku": value["sku"],
                    "product_name": value["display_name"] or key,
                    "normalized_product_name": key,
                    "qty": round(value["qty"], 2),
                    "sales": round(value["sales"], 2),
                    "gross_sales": round(value["gross_sales"], 2),
                    "net_sales": round(value["net_sales"], 2),
                    "supply_amount": round(value["supply_amount"], 2),
                    "delivery_amount": round(value["delivery_amount"], 2),
                    "channel_count": len(value["channels"]),
                }
                for key, value in product_totals.items()
            ],
            key=lambda item: (-item["qty"], item["product_name"]),
        ),
        "channel_product_sales": sorted(
            [
                {
                    "vendor_name": key[0],
                    "revenue_basis": channel_metadata.get(key[0], {}).get("revenue_basis", value["revenue_basis"]),
                    "date_basis": channel_metadata.get(key[0], {}).get("date_basis", value["date_basis"]),
                    "product_name": product_names.get(key[1], key[1]),
                    "normalized_product_name": key[1],
                    "qty": round(value["qty"], 2),
                    "sales": round(value["sales"], 2),
                    "gross_sales": round(value["gross_sales"], 2),
                    "net_sales": round(value["net_sales"], 2),
                    "supply_amount": round(value["supply_amount"], 2),
                    "delivery_amount": round(value["delivery_amount"], 2),
                }
                for key, value in channel_product.items()
            ],
            key=lambda item: (item["vendor_name"], -item["sales"], item["product_name"]),
        ),
        "daily_product_sales": sorted(
            [
                {
                    "business_date": key[0],
                    "product_name": product_names.get(key[1], key[1]),
                    "normalized_product_name": key[1],
                    "qty": round(value["qty"], 2),
                    "sales": round(value["sales"], 2),
                    "gross_sales": round(value["gross_sales"], 2),
                    "net_sales": round(value["net_sales"], 2),
                    "supply_amount": round(value["supply_amount"], 2),
                    "delivery_amount": round(value["delivery_amount"], 2),
                }
                for key, value in daily_product.items()
            ],
            key=lambda item: (item["business_date"], -item["sales"], item["product_name"]),
        ),
        "channel_definitions": sorted(channel_metadata.values(), key=lambda item: item["vendor_name"]),
    }


def build_summary(aggregates: dict[str, list[dict[str, Any]]], analyses: list[dict[str, Any]]) -> dict[str, Any]:
    total_sales = round(sum(item["sales"] for item in aggregates["product_sales"]), 2)
    total_qty = round(sum(item["qty"] for item in aggregates["product_qty"]), 2)
    channels = sorted({item["vendor_name"] for item in aggregates["channel_product_sales"] if item["vendor_name"]})
    dates = sorted({item["business_date"] for item in aggregates["daily_channel_sales"] if item["business_date"]})
    basis_rows = aggregates.get("channel_totals", [])
    revenue_bases = sorted({item.get("revenue_basis", "") for item in basis_rows if item.get("revenue_basis")})
    date_bases = sorted({item.get("date_basis", "") for item in basis_rows if item.get("date_basis")})
    return {
        "total_sales": total_sales,
        "total_qty": total_qty,
        "channel_count": len(channels),
        "date_count": len(dates),
        "total_gross_sales": round(sum(item.get("gross_sales", 0) or 0 for item in aggregates["product_sales"]), 2),
        "total_net_sales": round(sum(item.get("net_sales", 0) or 0 for item in aggregates["product_sales"]), 2),
        "total_supply_amount": round(sum(item.get("supply_amount", 0) or 0 for item in aggregates["product_sales"]), 2),
        "total_delivery_amount": round(sum(item.get("delivery_amount", 0) or 0 for item in aggregates["product_sales"]), 2),
        "revenue_basis_count": len(revenue_bases),
        "date_basis_count": len(date_bases),
        "revenue_bases": revenue_bases,
        "date_bases": date_bases,
        "top_channels": top_rows(aggregates["channel_totals"], key="sales", label="vendor_name", limit=5),
        "top_products_by_sales": top_rows(aggregates["product_sales"], key="sales", label="product_name", limit=10),
        "top_products_by_qty": top_rows(aggregates["product_qty"], key="qty", label="product_name", limit=10),
        "channel_definitions": basis_rows,
        "source_files": len(analyses),
    }


def export_report_workbook(path: Path, report: dict[str, Any]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    add_sheet(
        workbook,
        "Overview",
        [
            {"metric": "label", "value": report["label"]},
            {"metric": "source_count", "value": report["source_count"]},
            {"metric": "analysis_count", "value": report["analysis_count"]},
            {"metric": "record_count", "value": report["record_count"]},
            {"metric": "total_sales", "value": report["summary"]["total_sales"]},
            {"metric": "total_qty", "value": report["summary"]["total_qty"]},
            {"metric": "total_gross_sales", "value": report["summary"]["total_gross_sales"]},
            {"metric": "total_net_sales", "value": report["summary"]["total_net_sales"]},
            {"metric": "total_supply_amount", "value": report["summary"]["total_supply_amount"]},
            {"metric": "total_delivery_amount", "value": report["summary"]["total_delivery_amount"]},
            {"metric": "channel_count", "value": report["summary"]["channel_count"]},
            {"metric": "date_count", "value": report["summary"]["date_count"]},
            {"metric": "revenue_basis_count", "value": report["summary"]["revenue_basis_count"]},
            {"metric": "date_basis_count", "value": report["summary"]["date_basis_count"]},
        ],
    )
    add_sheet(
        workbook,
        "Standards",
        [
            {"metric": "channel_output_contract_id", "value": CHANNEL_OUTPUT_CONTRACT_ID},
            {"metric": "excel_postprocess_ruleset_id", "value": EXCEL_POSTPROCESS_RULESET_ID},
            {"metric": "product_analysis_master_schema_id", "value": PRODUCT_ANALYSIS_MASTER_SCHEMA_ID},
            {"metric": "channel_operating_model_id", "value": report["standards"]["channel_operating_model"]["id"]},
            {"metric": "group_by", "value": "normalized_product_name"},
        ],
    )
    add_sheet(
        workbook,
        "SourceFiles",
        [
            {
                "business_date": item.get("business_date", ""),
                "vendor_name": item.get("vendor_name", ""),
                "path": item.get("file", item.get("path", "")),
                "row_count": item.get("row_count", 0),
                "product_count": item.get("product_count", 0),
                "revenue_basis": item.get("channel_summary", {}).get("revenue_basis", item.get("operation_profile", {}).get("revenue_basis", "")),
                "date_basis": item.get("channel_summary", {}).get("date_basis", item.get("operation_profile", {}).get("date_basis", "")),
                "sales": item.get("totals", {}).get("sales", 0),
                "qty": item.get("totals", {}).get("qty", 0),
            }
            for item in report["analyses"]
        ],
    )
    add_sheet(workbook, "ChannelDefinitions", report["summary"]["channel_definitions"])
    add_sheet(workbook, "ChannelTotals", report["aggregates"]["channel_totals"])
    add_sheet(workbook, "DailyChannelSales", report["aggregates"]["daily_channel_sales"])
    add_sheet(workbook, "MonthlyChannelSales", report["aggregates"]["monthly_channel_sales"])
    add_sheet(workbook, "ProductSales", report["aggregates"]["product_sales"])
    add_sheet(workbook, "ProductQty", report["aggregates"]["product_qty"])
    add_sheet(workbook, "ChannelProductSales", report["aggregates"]["channel_product_sales"])
    add_sheet(workbook, "DailyProductSales", report["aggregates"]["daily_product_sales"])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def add_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet(title)
    if not rows:
        sheet["A1"] = "no data"
        return
    headers = list(rows[0].keys())
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    for row_index, row in enumerate(rows, start=2):
        for column_index, header in enumerate(headers, start=1):
            value = row.get(header, "")
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            if header in {"sales", "qty", "total_sales", "total_qty", "gross_sales", "net_sales", "supply_amount", "delivery_amount", "value"} and isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(12, max_length + 2), 48)


def build_summary_markdown(report: dict[str, Any]) -> str:
    summary = report["summary"]
    lines = [
        f"# 매출 리포트 {report['label']}",
        "",
        f"- 소스 파일 수: {report['source_count']}",
        f"- 분석 건수: {report['analysis_count']}",
        f"- 레코드 수: {report['record_count']}",
        f"- 총 매출액: {summary['total_sales']:.2f}",
        f"- 총 판매량: {summary['total_qty']:.2f}",
        f"- 총 주문/총매출 기준 합계: {summary['total_gross_sales']:.2f}",
        f"- 총 순매출 기준 합계: {summary['total_net_sales']:.2f}",
        f"- 총 공급/입고 기준 합계: {summary['total_supply_amount']:.2f}",
        f"- 총 납품 기준 합계: {summary['total_delivery_amount']:.2f}",
        f"- 채널 수: {summary['channel_count']}",
        f"- 집계 일수: {summary['date_count']}",
        f"- 매출 기준 종류 수: {summary['revenue_basis_count']}",
        f"- 기준일 종류 수: {summary['date_basis_count']}",
        "",
        "## 집계 기준 요약",
    ]
    for item in summary["channel_definitions"]:
        lines.append(
            f"- {item['vendor_name']}: 매출기준={item.get('revenue_basis', '') or '미정'} / 기준일={item.get('date_basis', '') or '미정'} / 수집방식={item.get('collection_mode', '') or '미정'}"
        )
    lines.extend(
        [
            "",
            "## 상위 채널",
        ]
    )
    for item in summary["top_channels"]:
        basis = next((row.get("revenue_basis", "") for row in summary["channel_definitions"] if row.get("vendor_name") == item["label"]), "")
        lines.append(f"- {item['label']}: {item['value']:.2f}" + (f" ({basis})" if basis else ""))
    lines.extend(["", "## 상위 품목(매출액)"])
    for item in summary["top_products_by_sales"]:
        lines.append(f"- {item['label']}: {item['value']:.2f}")
    lines.extend(["", "## 상위 품목(판매량)"])
    for item in summary["top_products_by_qty"]:
        lines.append(f"- {item['label']}: {item['value']:.2f}")
    lines.extend(
        [
            "",
            "## 첨부",
            "- 엑셀 리포트: 일별/월별 채널 매출, 품목별 매출, 품목별 판매량, 채널별 품목 매출 포함",
            "- 요약 문서: 상위 채널/품목, 전체 합계 포함",
            "- 채널 정의 시트: 매출 기준, 기준일, 수집방식, 검증방식 포함",
            "- 품목 집계 기준: normalized_product_name 마스터 스키마 적용",
        ]
    )
    return "\n".join(lines) + "\n"


def _new_amount_bucket() -> dict[str, Any]:
    return {
        "qty": 0.0,
        "sales": 0.0,
        "gross_sales": 0.0,
        "net_sales": 0.0,
        "supply_amount": 0.0,
        "delivery_amount": 0.0,
        "revenue_basis": "",
        "date_basis": "",
    }


def update_amount_bucket(
    bucket: dict[str, Any],
    qty: float,
    sales: float,
    gross_sales: float,
    net_sales: float,
    supply_amount: float,
    delivery_amount: float,
    revenue_basis: str,
    date_basis: str,
) -> None:
    bucket["qty"] += qty
    bucket["sales"] += sales
    bucket["gross_sales"] += gross_sales
    bucket["net_sales"] += net_sales
    bucket["supply_amount"] += supply_amount
    bucket["delivery_amount"] += delivery_amount
    if not bucket.get("revenue_basis") and revenue_basis:
        bucket["revenue_basis"] = revenue_basis
    if not bucket.get("date_basis") and date_basis:
        bucket["date_basis"] = date_basis


def create_email_draft(
    output_dir: Path,
    subject: str,
    to_addrs: list[str],
    cc_addrs: list[str],
    summary_markdown: str,
    attachments: list[Path],
    from_addr: str,
) -> Path:
    message = EmailMessage()
    message["Subject"] = subject
    if from_addr:
        message["From"] = from_addr
    if to_addrs:
        message["To"] = ", ".join(to_addrs)
    if cc_addrs:
        message["Cc"] = ", ".join(cc_addrs)
    message.set_content(summary_markdown)
    for attachment in attachments:
        mime_type, _ = mimetypes.guess_type(str(attachment))
        maintype, subtype = (mime_type or "application/octet-stream").split("/", 1)
        message.add_attachment(attachment.read_bytes(), maintype=maintype, subtype=subtype, filename=attachment.name)
    draft_path = output_dir / "report_email_draft.eml"
    draft_path.write_bytes(message.as_bytes())
    return draft_path


def send_email_bundle(
    secrets: SecretStore,
    smtp_profile: str,
    subject: str,
    to_addrs: list[str],
    cc_addrs: list[str],
    summary_markdown: str,
    attachments: list[Path],
) -> tuple[bool, str]:
    profile_name, smtp_config = resolve_email_profile(secrets, smtp_profile)
    provider = resolve_email_provider(smtp_config)
    validation = validate_smtp_profile(secrets, smtp_profile)
    if not validation.get("ready", False):
        return False, f"missing_{provider}_fields:{','.join(validation.get('missing_fields', []))}"
    if not to_addrs and not cc_addrs:
        return False, "missing_recipients"

    if provider == "resend":
        return send_resend_email(
            config=smtp_config,
            subject=subject,
            to_addrs=to_addrs,
            cc_addrs=cc_addrs,
            summary_markdown=summary_markdown,
            attachments=attachments,
        )

    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = smtp_config["from_addr"]
    if to_addrs:
        message["To"] = ", ".join(to_addrs)
    if cc_addrs:
        message["Cc"] = ", ".join(cc_addrs)
    message.set_content(summary_markdown)
    for attachment in attachments:
        mime_type, _ = mimetypes.guess_type(str(attachment))
        maintype, subtype = (mime_type or "application/octet-stream").split("/", 1)
        message.add_attachment(attachment.read_bytes(), maintype=maintype, subtype=subtype, filename=attachment.name)

    host = str(smtp_config["host"])
    port = int(smtp_config["port"])
    use_ssl = bool(smtp_config.get("use_ssl", False))
    use_tls = bool(smtp_config.get("use_tls", True))
    server_cls = smtplib.SMTP_SSL if use_ssl else smtplib.SMTP
    with server_cls(host, port, timeout=30) as server:
        if use_tls and not use_ssl:
            server.starttls()
        server.login(str(smtp_config["username"]), str(smtp_config["password"]))
        server.send_message(message)
    return True, ""


def validate_smtp_profile(secrets: SecretStore, smtp_profile: str) -> dict[str, Any]:
    profile_name, smtp_config = resolve_email_profile(secrets, smtp_profile)
    provider = resolve_email_provider(smtp_config)
    if provider == "resend":
        required = ["api_key", "from_addr"]
    else:
        required = ["host", "port", "username", "password", "from_addr"]
    missing = [key for key in required if not smtp_config.get(key)]
    return {
        "profile": profile_name,
        "provider": provider,
        "configured": bool(smtp_config),
        "ready": not missing,
        "missing_fields": missing,
    }


def resolve_email_profile(secrets: SecretStore, preferred_profile: str) -> tuple[str, dict[str, Any]]:
    candidates = [preferred_profile]
    for fallback in ("email", "resend", "smtp"):
        if fallback not in candidates:
            candidates.append(fallback)
    for candidate in candidates:
        config = secrets.get(candidate)
        if config:
            return candidate, config
    return preferred_profile, {}


def resolve_email_provider(config: dict[str, Any]) -> str:
    provider = str(config.get("provider", "") or "").strip().lower()
    if provider:
        return provider
    if config.get("api_key"):
        return "resend"
    return "smtp"


def send_resend_email(
    config: dict[str, Any],
    subject: str,
    to_addrs: list[str],
    cc_addrs: list[str],
    summary_markdown: str,
    attachments: list[Path],
) -> tuple[bool, str]:
    payload: dict[str, Any] = {
        "from": str(config["from_addr"]),
        "to": to_addrs,
        "subject": subject,
        "text": summary_markdown,
    }
    if cc_addrs:
        payload["cc"] = cc_addrs
    reply_to = config.get("reply_to")
    if reply_to:
        payload["replyTo"] = reply_to
    if attachments:
        payload["attachments"] = [
            {
                "filename": attachment.name,
                "content": base64.b64encode(attachment.read_bytes()).decode("ascii"),
            }
            for attachment in attachments
        ]

    base_url = str(config.get("base_url", "https://api.resend.com")).rstrip("/")
    request = urlrequest.Request(
        url=f"{base_url}/emails",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {config['api_key']}",
            "Content-Type": "application/json",
            "User-Agent": "openclaw-sales-pipeline/1.0",
        },
        method="POST",
    )
    max_attempts = int(config.get("max_attempts", 3) or 3)
    last_error = ""
    for attempt in range(1, max_attempts + 1):
        try:
            with urlrequest.urlopen(request, timeout=30) as response:
                response_body = response.read().decode("utf-8")
            parsed = json.loads(response_body or "{}")
            if parsed.get("id"):
                return True, ""
            last_error = f"resend_unexpected_response:{response_body[:240]}"
            if attempt < max_attempts:
                time.sleep(resend_backoff_seconds(attempt))
                continue
            return False, last_error
        except urlerror.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace")
            last_error = f"resend_http_error:{exc.code}:{body[:240]}"
            if attempt < max_attempts and is_retryable_resend_status(exc.code):
                time.sleep(resend_backoff_seconds(attempt, exc.headers.get("Retry-After")))
                continue
            return False, last_error
        except Exception as exc:
            last_error = f"resend_request_failed:{type(exc).__name__}:{str(exc)[:240]}"
            if attempt < max_attempts:
                time.sleep(resend_backoff_seconds(attempt))
                continue
            return False, last_error
    return False, last_error or "resend_unknown_failure"


def is_retryable_resend_status(status_code: int) -> bool:
    return status_code in {408, 409, 425, 429, 500, 502, 503, 504}


def resend_backoff_seconds(attempt: int, retry_after: str | None = None) -> float:
    if retry_after:
        try:
            return max(0.5, min(float(retry_after), 30.0))
        except ValueError:
            pass
    return min(2 ** (attempt - 1), 8)


def top_rows(rows: list[dict[str, Any]], key: str, label: str, limit: int) -> list[dict[str, Any]]:
    return [{"label": item[label], "value": item[key]} for item in rows[:limit]]


def sort_rows(rows: list[dict[str, Any]], keys: list[str]) -> list[dict[str, Any]]:
    return sorted(rows, key=lambda item: tuple(item.get(key, "") for key in keys))


def looks_like_date(value: str) -> bool:
    try:
        datetime.strptime(value, "%Y-%m-%d")
        return True
    except ValueError:
        return False


def infer_label(analyses: list[dict[str, Any]], date_from: str | None, date_to: str | None) -> str:
    if date_from and date_to and date_from != date_to:
        return f"{date_from}_to_{date_to}"
    if date_from:
        return date_from
    dates = sorted({item.get("business_date", "") for item in analyses if item.get("business_date")})
    if len(dates) == 1:
        return dates[0]
    if dates:
        return f"{dates[0]}_to_{dates[-1]}"
    return datetime.now().strftime("%Y-%m-%d")
